import pandas as pd
import time
import logging
import argparse

from openpyxl import load_workbook

from utilities.get_CMD_EXEC import compare_CMD_EXEC
from utilities.get_FILE_CHECK_NOT import compare_FILE_CHECK_NOT
from utilities.get_FILE_CHECK import compare_FILE_CHECK
from utilities.get_FILE_CONTENT_CHECK_NOT import compare_FILE_CONTENT_CHECK_NOT
from utilities.get_FILE_CONTENT_CHECK import compare_FILE_CONTENT_CHECK


def get_actual_values(data_dict: dict) -> dict:
    '''
    This function takes a dictionary of data as input. For each type of audit, 
    it calls the appropriate function to compare the actual and expected values. 
    The function then returns a new dictionary with the comparison results.

    :param data_dict: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    :return: 
        A new dictionary with the same keys as data_dict but the values replaced 
        with DataFrames that include the results of the comparison between the 
        actual and expected values.
    '''

    new_dict = {}

    for key in data_dict.keys():
        print(key)

        try:
            if key == "CMD_EXEC":
                new_df = compare_CMD_EXEC(data_dict)
            elif key == "FILE_CHECK_NOT":
                new_df = compare_FILE_CHECK_NOT(data_dict)
            elif key == "FILE_CHECK":
                new_df = compare_FILE_CHECK(data_dict)
            elif key == "FILE_CONTENT_CHECK_NOT":
                new_df = compare_FILE_CONTENT_CHECK_NOT(data_dict)
            elif key == "FILE_CONTENT_CHECK":
                new_df = compare_FILE_CONTENT_CHECK(data_dict)
            else:
                new_df = pd.DataFrame()

            new_dict[key] = new_df
        except Exception as e:
            print(e)

    return new_dict


def read_file(fname: str) -> dict:
    '''
    This function reads the provided audit file and returns a dictionary where 
    the keys are different audit types and the values are corresponding audit 
    data in DataFrame format.

    :param fname: 
        A string representing the filename of the audit file.
    :return: 
        A dictionary with keys representing different audit types and values as 
        DataFrames containing the audit data.
    '''

    data_dict = {
        "FILE_CHECK_NOT": [],
        "CMD_EXEC": [],
        "FILE_CONTENT_CHECK_NOT": [],
        "FILE_CHECK": [],
        "BANNER_CHECK": [],
        "FILE_CONTENT_CHECK": []
    }

    xl = pd.ExcelFile(fname)

    # df = xl.parse(sheet_name=0)

    for ptype in data_dict:
        try:
            df0 = xl.parse(sheet_name=ptype, dtype={'Mask': str})
            data_dict[ptype] = df0.applymap(remove_illegal_chars)
        except ValueError as e:
            logging.error(f"{ptype} not found")

    return data_dict


def remove_illegal_chars(val: str) -> str:
    '''
    This function checks if the given value is a string, and if so, removes any 
    non-printable characters.

    :param val: 
        A string potentially containing non-printable characters.
    :return: 
        The same string but with any non-printable characters removed.
    '''

    if isinstance(val, str):
        # Remove control characters
        val = ''.join(ch for ch in val if ch.isprintable())
    return val


def save_file(out_fname: str, data_dict_list: list, ip_addr: str) -> None:
    ''' 
    This function processes the comparison results and saves them into an Excel file.

    :param out_fname: 
        A string representing the filename of the output file.
    :param data_dict_list: 
        A list of dictionaries containing comparison results.
    :param ip_addr: 
        A string representing an IP address.
    :return: 
        None
    '''

    if data_dict_list == []:
        return

    all_frames = []

    for data_dict in data_dict_list:
        frames = []

        for df in data_dict.values():
            frames.append(df)
        result_rowwise = pd.concat(frames)
        all_frames.append(result_rowwise.reset_index(drop=True))

    result = pd.concat(all_frames, axis=1)

    result = result.loc[:, ~result.columns.duplicated()]

    column_names = result.columns.tolist()

    value_n_result = column_names[15:]
    print(column_names[0])
    ip_list = [ip_addr, '']
    name_list = []

    for i in range(len(value_n_result)):
        if i % 2 == 0:
            # ip = value_n_result[i].split('|')[0].strip()
            # ip_list.append(ip)
            name_list.append('Actual Value')
        else:
            # ip_list.append('')
            name_list.append('Result')

    new_data = ['Checklist', 'Type', 'Index', 'Description', 'Solution',
                'File',  'Owner', 'Mask', 'Required', 'Group', 'CMD', 'Expect', 'Regex', 'Content', 'Is_substring'] + name_list
    result.columns = ['Checklist', 'Type', 'Index', 'Description', 'Solution',
                      'File',  'Owner', 'Mask', 'Required', 'Group', 'CMD', 'Expect', 'Regex', 'Content', 'Is_substring'] + ip_list

    new_df = pd.DataFrame(
        [new_data + [''] * (result.shape[1] - len(new_data))], columns=result.columns)
    result = pd.concat([new_df, result]).reset_index(drop=True)

    # Apply the function to each string column in the DataFrame
    result = result.applymap(remove_illegal_chars)

    # Save DataFrame to a new Excel file
    result.to_excel(out_fname, index=False)

    # Load the workbook and select the sheet
    wb = load_workbook(out_fname)
    ws = wb.active

    # Merge the appropriate cells in the new first row
    for col in range(1, 16):  # adjust these values as needed
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=2, end_column=col)

    for ip_col in range(16, len(result.columns)):  # adjust these values as needed
        if ip_col % 2 == 0:
            ws.merge_cells(start_row=1, start_column=ip_col,
                           end_row=1, end_column=ip_col+1)
        else:
            continue

    # Save the workbook
    wb.save(out_fname)
    print((f"Result saved into {out_fname}"))


'''
This is the main function that gets executed when the script runs. It parses 
command-line arguments, reads the audit file, compares actual and expected 
values for each audit type, and finally saves the comparison results into an 
output file.
'''
if __name__ == '__main__':

    # my_parser = argparse.ArgumentParser(
    #     description='A Security Audit Program')

    # # Add the arguments
    # my_parser.add_argument('--ps_result',
    #                        type=str,
    #                        required=True,
    #                        help='The path of result file after running the PowerShell script')

    # my_parser.add_argument('--output',
    #                        type=str,
    #                        required=True,
    #                        help='The path of output file')

    # my_parser.add_argument('--audit',
    #                        type=str,
    #                        required=True,
    #                        help='The path of audit file')

    # # Execute parse_args()
    # args = my_parser.parse_args()

    # print('PowerShell result file:', args.ps_result)
    # print('Output file:', args.output)
    # print('Audit file:', args.audit)

    start_t0 = time.time()

    result_fname = "output_Debian11_su.txt"
    # result_fname = "output_num.txt"
    # result_fname = args.ps_result

    output_list = []
    with open(result_fname, 'r') as file:
        lines = file.readlines()
    # exit()

    single_line = '\n'.join(line.strip() for line in lines)

    # actual value list
    output_list = single_line.strip().split("==|==")
    output_list.pop(0)

    # print(output_list)
    # exit()

    audit_fname = "src\Audit\CIS_Debian_Linux_10_v2.0.0_L1_Server.xlsx"
    # audit_fname = args.audit

    data_dict = read_file(audit_fname)

    # add actual value to the audit file
    head = 0
    for key in data_dict:
        if key == "CMD_EXEC" or key == "FILE_CHECK_NOT" or key == "FILE_CHECK" or key == "FILE_CONTENT_CHECK_NOT" or key == "FILE_CONTENT_CHECK":
            length = len(data_dict[key])
            data_dict[key]['Actual Value'] = output_list[head: (head+length)]
            head += length

    new_dict = get_actual_values(data_dict)

    results = []
    results.append(new_dict)

    ip_addr = "IP"
    # # write output file
    save_file("debian11_result_su2.xlsx", results, ip_addr)
